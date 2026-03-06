"""
데이터 로더 모듈 (Data-Driven Testing)

JSON, CSV, Excel 파일을 통합 로드하고,
생성 스크립트에 삽입할 로더 코드를 반환합니다.
"""

import os
import json
import csv
from typing import List, Dict, Optional
from utils.logger import setup_logger

logger = setup_logger(__name__)


class DataLoader:
    """
    통합 데이터 로더: JSON / CSV / Excel 파일 지원

    Usage:
        loader = DataLoader()
        data = loader.load("data/test_cases.json")   # -> List[Dict]
        code = loader.generate_loader_code("data/test_cases.json")  # -> str (Python code)
    """

    SUPPORTED_FORMATS = {"json", "csv", "xlsx", "xls"}

    # ================================================================
    # 포맷 감지
    # ================================================================

    @staticmethod
    def detect_format(file_path: str) -> str:
        """
        파일 확장자로 데이터 포맷 감지

        Args:
            file_path: 데이터 파일 경로

        Returns:
            "json" | "csv" | "excel"

        Raises:
            ValueError: 지원하지 않는 확장자
        """
        ext = os.path.splitext(file_path)[1].lower().lstrip(".")
        if ext == "json":
            return "json"
        elif ext == "csv":
            return "csv"
        elif ext in ("xlsx", "xls"):
            return "excel"
        else:
            raise ValueError(f"지원하지 않는 데이터 파일 형식: .{ext} (지원: .json, .csv, .xlsx, .xls)")

    # ================================================================
    # 데이터 로드
    # ================================================================

    def load(self, file_path: str) -> List[Dict]:
        """
        파일에서 테스트 데이터 로드

        Args:
            file_path: 데이터 파일 경로

        Returns:
            List[Dict]: 행 단위 딕셔너리 리스트
        """
        if not os.path.exists(file_path):
            logger.error(f"데이터 파일 없음: {file_path}")
            return []

        fmt = self.detect_format(file_path)
        logger.info(f"데이터 로드 중: {file_path} (format={fmt})")

        if fmt == "json":
            return self._load_json(file_path)
        elif fmt == "csv":
            return self._load_csv(file_path)
        elif fmt == "excel":
            return self._load_excel(file_path)
        return []

    def _load_json(self, file_path: str) -> List[Dict]:
        """JSON 파일 로드"""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            # JSON이 리스트가 아닌 경우 처리
            if isinstance(data, dict):
                # {"test_cases": [...]} 형태 지원
                for key in ("test_cases", "data", "rows", "cases"):
                    if key in data and isinstance(data[key], list):
                        data = data[key]
                        break
                else:
                    data = [data]  # 단일 객체 → 리스트로 감싸기

            if not isinstance(data, list):
                logger.error(f"JSON 구조 오류: 리스트 또는 딕셔너리 기대")
                return []

            # 모든 값을 문자열로 정규화
            result = []
            for row in data:
                if isinstance(row, dict):
                    result.append({str(k): str(v) if v is not None else "" for k, v in row.items()})
            
            logger.info(f"JSON 로드 완료: {len(result)}행")
            return result

        except json.JSONDecodeError as e:
            logger.error(f"JSON 파싱 실패: {e}")
            return []
        except Exception as e:
            logger.error(f"JSON 로드 실패: {e}")
            return []

    def _load_csv(self, file_path: str) -> List[Dict]:
        """CSV 파일 로드"""
        try:
            result = []
            with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # 값 정규화 (None → 빈 문자열)
                    cleaned = {str(k).strip(): str(v).strip() if v else "" for k, v in row.items()}
                    result.append(cleaned)

            logger.info(f"CSV 로드 완료: {len(result)}행")
            return result

        except Exception as e:
            logger.error(f"CSV 로드 실패: {e}")
            return []

    def _load_excel(self, file_path: str) -> List[Dict]:
        """Excel 파일 로드"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active

            rows = sheet.iter_rows(values_only=True)
            try:
                headers = next(rows)
                headers = [str(h).strip() for h in headers if h is not None]
            except StopIteration:
                logger.warning("Excel 데이터 없음")
                return []

            result = []
            for row in rows:
                row_data = {}
                for i, h in enumerate(headers):
                    if i < len(row):
                        val = row[i]
                        row_data[h] = str(val) if val is not None else ""
                    else:
                        row_data[h] = ""
                result.append(row_data)

            wb.close()
            logger.info(f"Excel 로드 완료: {len(result)}행")
            return result

        except Exception as e:
            logger.error(f"Excel 로드 실패: {e}")
            return []

    # ================================================================
    # 생성 스크립트용 로더 코드 반환
    # ================================================================

    def generate_loader_code(self, file_path: str) -> str:
        """
        생성될 pytest 스크립트에 삽입할 데이터 로더 Python 코드 반환

        Args:
            file_path: 데이터 파일 경로

        Returns:
            str: Python 코드 문자열 (포맷별 로더 함수)
        """
        fmt = self.detect_format(file_path)
        safe_path = file_path.replace("\\", "/")

        if fmt == "json":
            return self._gen_json_loader_code(safe_path)
        elif fmt == "csv":
            return self._gen_csv_loader_code(safe_path)
        elif fmt == "excel":
            return self._gen_excel_loader_code(safe_path)
        return ""

    def _gen_json_loader_code(self, safe_path: str) -> str:
        return f"""
import json
import sys
import os

def get_test_data():
    file_path = r"{safe_path}"
    print(f"\\\\n[INFO] JSON 데이터 로드 중: {{file_path}}")
    if not os.path.exists(file_path):
        print(f"[ERROR] 파일 없음: {{file_path}}")
        return []
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        # 리스트 또는 딕셔너리 처리
        if isinstance(data, dict):
            for key in ("test_cases", "data", "rows", "cases"):
                if key in data and isinstance(data[key], list):
                    data = data[key]
                    break
            else:
                data = [data]
        if not isinstance(data, list):
            print("[ERROR] JSON 구조 오류")
            return []
        result = []
        for row in data:
            if isinstance(row, dict):
                result.append({{str(k): str(v) if v is not None else "" for k, v in row.items()}})
        if not result: print("[WARN] 데이터 없음")
        return result
    except Exception as e:
        print(f"\\\\n[FATAL] JSON 읽기 실패: {{e}}")
        return []
"""

    def _gen_csv_loader_code(self, safe_path: str) -> str:
        return f"""
import csv
import sys
import os

def get_test_data():
    file_path = r"{safe_path}"
    print(f"\\\\n[INFO] CSV 데이터 로드 중: {{file_path}}")
    if not os.path.exists(file_path):
        print(f"[ERROR] 파일 없음: {{file_path}}")
        return []
    try:
        result = []
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                cleaned = {{str(k).strip(): str(v).strip() if v else "" for k, v in row.items()}}
                result.append(cleaned)
        if not result: print("[WARN] 데이터 없음")
        return result
    except Exception as e:
        print(f"\\\\n[FATAL] CSV 읽기 실패: {{e}}")
        return []
"""

    def _gen_excel_loader_code(self, safe_path: str) -> str:
        return f"""
import openpyxl
import sys
import os

def get_test_data():
    file_path = r"{safe_path}"
    print(f"\\\\n[INFO] 엑셀 로드 중: {{file_path}}")
    if not os.path.exists(file_path):
        print(f"[ERROR] 파일 없음: {{file_path}}")
        return []
    try:
        data = []
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        # 헤더 읽기
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
            headers = [str(h).strip() for h in headers if h is not None]
        except StopIteration:
            print("[WARN] 데이터 없음")
            return []

        # 데이터 읽기
        for row in rows:
            row_data = {{}}
            for i, h in enumerate(headers):
                if i < len(row):
                    val = row[i]
                    row_data[h] = str(val) if val is not None else ""
                else:
                    row_data[h] = ""
            data.append(row_data)

        if not data: print("[WARN] 데이터 없음")
        return data
    except Exception as e:
        print(f"\\\\n[FATAL] 엑셀 읽기 실패: {{e}}")
        return []
"""
